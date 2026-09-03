"""Extract the spending inputs from the ONS Family Spending workbook.

Regenerates `src/iran_impact/inputs/ons_family_spending_a6_fye2024.csv` from
the published workbook, so the numbers the model uses can be traced to a
downloaded file rather than transcribed by hand.

    python scripts/extract_ons_a6.py

Requires `openpyxl` and network access. The workbook is not committed; the
extracted CSV is, along with the SHA-256 of the workbook it came from.
"""

import argparse
import csv
import hashlib
import urllib.request
from pathlib import Path

SOURCE_URL = (
    "https://www.ons.gov.uk/file?uri=%2Fpeoplepopulationandcommunity"
    "%2Fpersonalandhouseholdfinances%2Fexpenditure%2Fdatasets"
    "%2Ffamilyspendingworkbook1detailedexpenditureandtrends%2Ffye2024"
    "%2Fworkbook1detailedexpenditureandtrends.xlsx"
)
BULLETIN_URL = (
    "https://www.ons.gov.uk/peoplepopulationandcommunity"
    "/personalandhouseholdfinances/expenditure/bulletins/familyspendingintheuk"
    "/april2023tomarch2024"
)

SHEET = "A6"
# Table A6 lays the ten gross-income decile groups out in columns E-N, with
# the all-households column in O.
FIRST_DECILE_COLUMN = 5
ALL_HOUSEHOLDS_COLUMN = 15

# The two commodity lines the model uses, by their A6 classification code.
COMMODITIES = {
    "7.2.2": "transport_fuel",
    "1": "food_and_non_alcoholic_drinks",
}

OUTPUT = (
    Path(__file__).resolve().parents[1]
    / "src"
    / "iran_impact"
    / "inputs"
    / "ons_family_spending_a6_fye2024.csv"
)


def _cell(sheet, row, column):
    """A6 stores its figures as text, e.g. '19.80'."""
    value = sheet.cell(row, column).value
    if value is None:
        return None
    return float(str(value).replace(",", ""))


def _find_row(sheet, code, label_column):
    for row in range(1, sheet.max_row + 1):
        if str(sheet.cell(row, label_column).value or "").strip() == code:
            return row
    raise LookupError(f"no row with code {code!r} in column {label_column}")


def extract(workbook_path):
    import openpyxl

    sheet = openpyxl.load_workbook(workbook_path, data_only=True)[SHEET]
    rows = []
    for code, name in COMMODITIES.items():
        # A6 indents its classification codes: top-level codes sit in the
        # first column and nested ones move right, so search the first three.
        for label_column in (1, 2, 3):
            try:
                row = _find_row(sheet, code, label_column)
            except LookupError:
                continue
            break
        else:
            raise LookupError(f"code {code!r} not found in {SHEET}")

        # The description is the cell immediately right of the code.
        description = str(sheet.cell(row, label_column + 1).value or "").strip()
        all_households = _cell(sheet, row, ALL_HOUSEHOLDS_COLUMN)
        for offset in range(10):
            weekly = _cell(sheet, row, FIRST_DECILE_COLUMN + offset)
            rows.append(
                {
                    "commodity": name,
                    "a6_code": code,
                    "a6_description": description,
                    "gross_income_decile": offset + 1,
                    "weekly_spend_gbp": f"{weekly:.2f}",
                    "all_households_weekly_spend_gbp": f"{all_households:.2f}",
                }
            )
    return rows


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--workbook",
        help="path to an already-downloaded workbook; downloaded if omitted",
    )
    args = parser.parse_args()

    if args.workbook:
        path = Path(args.workbook)
        payload = path.read_bytes()
    else:
        print(f"downloading {SOURCE_URL}")
        with urllib.request.urlopen(SOURCE_URL) as response:
            payload = response.read()
        path = Path("workbook1detailedexpenditureandtrends.xlsx")
        path.write_bytes(payload)

    digest = hashlib.sha256(payload).hexdigest()
    rows = extract(path)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT.open("w", newline="") as handle:
        handle.write(
            "# ONS Family Spending in the UK, financial year ending 2024.\n"
            f"# Table: {SHEET} - detailed household expenditure by gross "
            "income decile group, UK\n"
            "# Units: mean GBP per week per household, in FYE 2024 prices\n"
            "# Grouping variable: gross household income decile (ONS's own "
            "grouping for this table)\n"
            f"# Bulletin: {BULLETIN_URL}\n"
            f"# Workbook: {SOURCE_URL}\n"
            f"# Workbook SHA-256: {digest}\n"
            "# Regenerate with: python scripts/extract_ons_a6.py\n"
        )
        # Unix line endings, so `git diff --check` does not report every data
        # row as trailing whitespace (#12 review S2).
        writer = csv.DictWriter(
            handle, fieldnames=list(rows[0]), lineterminator="\n"
        )
        writer.writeheader()
        writer.writerows(rows)
    print(f"wrote {len(rows)} rows to {OUTPUT}")
    print(f"workbook sha256 {digest}")


if __name__ == "__main__":
    main()
